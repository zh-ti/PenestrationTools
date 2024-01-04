from whoisxmlapi_api import Api as Whoisxmlapi
from whoisxmlapi_reptile import Api as WhoisxmlReptile
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, NamedStyle

argParser = argparse.ArgumentParser("Search subdomain by domain")

argParser.add_argument('-d', '--domain', type=str, required=False, help='The domain name that needs to be query')
argParser.add_argument('-i', '--input-file', type=str, required=False,
                       help='The file collection of the domain name that needs to be query')
argParser.add_argument('-o', '--output-file', type=str, required=False, default="./subdomains.xlsx",
                       help='The output file of the query result')

args = argParser.parse_args()

# 第一行的样式
header_style = NamedStyle(name="header_style")
header_style.font = Font(bold=True)
header_style.fill = PatternFill(start_color="d3e3fd", end_color="d3e3fd", fill_type="solid")
# 列宽
Column_Width = 30

class Main:
    # 要查询的域名
    domains = set()
    # 查询到的子域名
    subdomains = {}

    def __init__(self):
        # 获取参数
        domain_str = args.domain
        input_file = args.input_file
        domain_str = 'zh-ti.top'
        if domain_str:
            if ',' in domain_str:
                self.domains.update(domain_str.split(','))
            else:
                self.domains.update([domain_str])
        elif input_file:
            with open(file=input_file, encoding='utf-8', buffering=1024) as file:
                self.domains.update(file.readlines())

    def run(self):
        # 将子域名按照域名分类
        for domain in self.domains:
            self.subdomains[domain] = set()
        # 初始化api
        whoisxmlapi = Whoisxmlapi(apiKey="at_weRRUsIZtIaaxGzvAql8WUc2XaJLk")
        whoisxmlReptile = WhoisxmlReptile()
        # 查询所有域名的子域
        for domain in self.domains:
            self.subdomains[domain].update(whoisxmlapi.search(domain))
            self.subdomains[domain].update(whoisxmlReptile.search(domain))

        self.export(args.output_file)

    def export(self, output_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        max_col = len(self.subdomains)
        max_row = 0

        for domain in self.subdomains:
            max_row = max(max_row, len(self.subdomains[domain]))

        for (index, col) in enumerate(sheet.iter_cols(min_row=1, max_row=max_row + 1, min_col=1, max_col=max_col)):
            if len(col) == 0:
                break
            domain, subdomains = self.subdomains.popitem()
            sheet.column_dimensions[get_column_letter(index + 1)].width = Column_Width
            for (subIndex, cell) in enumerate(col):
                if subIndex == 0:
                    cell.style = header_style
                    cell.value = f"{domain} ({len(subdomains)})"
                else:
                    cell.value = subdomains.pop()
                if len(subdomains) == 0:
                    break
        workbook.save(output_file)


Main().run()
